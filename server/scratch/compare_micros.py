import openpyxl

wb = openpyxl.load_workbook("Plan de entrenamiento  2.xlsx", data_only=True)
micro_sheets = [s for s in wb.sheetnames if s.startswith("MICRO")]

print("Analyzing microcycle exercises across sheets:")
for sheet_name in micro_sheets:
    sheet = wb[sheet_name]
    exercises = []
    current_day = ""
    for r in range(1, sheet.max_row + 1):
        val = sheet.cell(row=r, column=2).value
        if val:
            val_str = str(val).strip()
            if any(day_name in val_str for day_name in ["PUSH 1", "PULL 1", "LEGS 1", "PUSH 2", "PULL 2", "LEGS 2"]):
                current_day = val_str
            elif val_str not in ["EJERCICIO", "PUSH 1", "PULL 1", "LEGS 1", "PUSH 2", "PULL 2", "LEGS 2", "EJERCICIOS", "TECNICA", "NOTAS"]:
                exercises.append(f"{current_day}:{val_str}")
    print(f"Sheet {sheet_name} has {len(exercises)} exercise slots. First few: {exercises[:5]}")
