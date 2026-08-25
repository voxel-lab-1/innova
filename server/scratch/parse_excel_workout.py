import openpyxl
import json
import os
import re

file_path = "Plan de entrenamiento  2.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

# 1. Parse Alternatives sheet
alt_sheet = wb["ALTERNATIVAS"]
alternatives_map = {} # exercise_name -> list of alternatives

def clean_name_and_link(val, cell_hyperlink=None):
    if not val:
        return "", ""
    val_str = str(val).strip()
    
    # Try to extract URL from cell hyperlink if present
    link = ""
    if cell_hyperlink:
        link = cell_hyperlink.target
    
    # If not in cell_hyperlink, try regex extraction of URL inside parentheses
    if not link:
        url_match = re.search(r'\((https?://[^\)]+)\)', val_str)
        if url_match:
            link = url_match.group(1)
            
    # Clean exercise name (remove (http...) link parts)
    clean_name = re.sub(r'\s*\(https?://[^\)]+\)', '', val_str).strip()
    clean_name = re.sub(r'\s*\(Link:.*?\)', '', clean_name).strip()
    return clean_name, link

# Go through ALTERNATIVAS sheet rows
for r in range(1, alt_sheet.max_row + 1):
    ex_cell = alt_sheet.cell(row=r, column=2)
    alt1_cell = alt_sheet.cell(row=r, column=3)
    alt2_cell = alt_sheet.cell(row=r, column=4)
    
    ex_name, ex_link = clean_name_and_link(ex_cell.value, ex_cell.hyperlink)
    alt1_name, alt1_link = clean_name_and_link(alt1_cell.value, alt1_cell.hyperlink)
    alt2_name, alt2_link = clean_name_and_link(alt2_cell.value, alt2_cell.hyperlink)
    
    if ex_name and ex_name not in ["EJERCICIO", "PUSH 1", "PULL 1", "LEGS 1", "PUSH 2", "PULL 2", "LEGS 2"]:
        alts = []
        if alt1_name and alt1_name != "-":
            alts.append({"name": alt1_name, "videoUrl": alt1_link})
        if alt2_name and alt2_name != "-":
            alts.append({"name": alt2_name, "videoUrl": alt2_link})
        
        # Key on clean lowercased name for matching
        alternatives_map[ex_name.lower()] = alts

print(f"Parsed {len(alternatives_map)} exercise alternative records.")

# 2. Parse Workout Sheet (MICRO 1 INTRODUCCIÓN)
intro_sheet = wb["MICRO 1 INTRODUCCIÓN"]
workout_days = {} # day_name -> list of exercises
current_day = None

# We can define muscle groups map
muscle_group_map = {
    "push 1": "chest",
    "pull 1": "back",
    "legs 1": "legs",
    "push 2": "chest",
    "pull 2": "back",
    "legs 2": "legs"
}

for r in range(1, intro_sheet.max_row + 1):
    val_b = intro_sheet.cell(row=r, column=2).value
    val_b_str = str(val_b).strip() if val_b else ""
    
    if any(day_header in val_b_str for day_header in ["PUSH 1", "PULL 1", "LEGS 1", "PUSH 2", "PULL 2", "LEGS 2"]):
        current_day = val_b_str.lower()
        workout_days[current_day] = []
        continue
        
    if current_day and val_b and val_b_str not in ["EJERCICIO", "EJERCICIOS", "TECNICA", "NOTAS", "TÉCNICA"]:
        cell_ex = intro_sheet.cell(row=r, column=2)
        ex_name, ex_link = clean_name_and_link(cell_ex.value, cell_ex.hyperlink)
        
        technique = str(intro_sheet.cell(row=r, column=3).value or "").strip()
        notes = str(intro_sheet.cell(row=r, column=4).value or "").strip()
        
        try:
            sets = int(intro_sheet.cell(row=r, column=6).value or 3)
        except:
            sets = 3
            
        reps = str(intro_sheet.cell(row=r, column=7).value or "8-12").strip()
        rest = str(intro_sheet.cell(row=r, column=8).value or "2-3'").strip()
        rir = str(intro_sheet.cell(row=r, column=9).value or "2-3").strip()
        
        # Match alternatives
        ex_alts = alternatives_map.get(ex_name.lower(), [])
        
        # Deduced muscle group
        ex_muscle = muscle_group_map.get(current_day, "legs")
        # Overwrite specific muscle groups if name contains keywords
        name_lower = ex_name.lower()
        if "tríceps" in name_lower or "pushdown" in name_lower or "katana" in name_lower:
            ex_muscle = "arms"
        elif "bíceps" in name_lower or "curl" in name_lower:
            ex_muscle = "arms"
        elif "lateral" in name_lower or "press de hombro" in name_lower or "shoulder press" in name_lower:
            ex_muscle = "shoulders"
        elif "abdominal" in name_lower or "plancha" in name_lower:
            ex_muscle = "core"
            
        workout_days[current_day].append({
            "name": ex_name,
            "videoUrl": ex_link,
            "muscleGroup": ex_muscle,
            "sets": sets,
            "reps": reps,
            "rest": rest,
            "rir": rir,
            "technique": technique if technique != "None" else "",
            "notes": notes if notes != "None" else "",
            "alternatives": ex_alts
        })

# Save to templates.json
templates_data = {
    "source": "Plan de entrenamiento  2.xlsx",
    "days": workout_days
}

output_json = "server/trainingTemplates.json"
with open(output_json, "w", encoding="utf-8") as jf:
    json.dump(templates_data, jf, indent=2, ensure_ascii=False)

print("Parsed and saved training templates json successfully at:", output_json)
print("Days parsed:", list(workout_days.keys()))
for d, exs in workout_days.items():
    print(f" - {d}: {len(exs)} exercises")
