import openpyxl
import os

file_path = "Plan de entrenamiento  2.xlsx"
if not os.path.exists(file_path):
    print("File not found:", file_path)
    exit(1)

wb = openpyxl.load_workbook(file_path, data_only=False)

output_path = "server/scratch/training_plan_dump.txt"
with open(output_path, "w", encoding="utf-8") as f:
    f.write(f"Workbook Sheets: {wb.sheetnames}\n")
    
    for sheet_name in wb.sheetnames:
        f.write(f"\n==================================================\n")
        f.write(f"SHEET: {sheet_name}\n")
        f.write(f"==================================================\n")
        sheet = wb[sheet_name]
        
        # Determine max rows and cols
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        for r in range(1, max_row + 1):
            row_cells = []
            has_data = False
            for c in range(1, max_col + 1):
                cell = sheet.cell(row=r, column=c)
                val = cell.value
                
                # Check link
                link = ""
                if cell.hyperlink:
                    link = f" ({cell.hyperlink.target})"
                
                if val is not None:
                    has_data = True
                    row_cells.append(f"C{c}[{val}{link}]")
                else:
                    row_cells.append(f"C{c}[]")
            
            if has_data:
                # Remove trailing empty cells to make it clean
                while len(row_cells) > 0 and row_cells[-1] == f"C{len(row_cells)}[]":
                    row_cells.pop()
                f.write(f"R{r}: " + " | ".join(row_cells) + "\n")

print("Dump created successfully at:", output_path)
