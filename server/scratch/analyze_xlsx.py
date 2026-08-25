import openpyxl
import os

file_path = "Plan de entrenamiento  2.xlsx"
if not os.path.exists(file_path):
    print("File not found:", file_path)
    exit(1)

wb = openpyxl.load_workbook(file_path, data_only=False)
print("Sheet names:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    print(f"\n--- Analyzing Sheet: {sheet_name} ---")
    sheet = wb[sheet_name]
    print(f"Dimensions: {sheet.dimensions}")
    
    # Dump first 10 rows and check for hyperlinks
    for r in range(1, 40):
        row_values = []
        has_hyperlinks = False
        for c in range(1, 15):
            cell = sheet.cell(row=r, column=c)
            val = cell.value
            link = ""
            if cell.hyperlink:
                link = f" (Link: {cell.hyperlink.target})"
                has_hyperlinks = True
            
            if val is not None:
                row_values.append(f"col{c}:{val}{link}")
        
        if row_values:
            print(f"Row {r:02d}:", " | ".join(row_values))
